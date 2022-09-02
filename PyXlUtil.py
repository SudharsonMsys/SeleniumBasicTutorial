import openpyxl
from openpyxl.styles import PatternFill


def rowcount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.max_row


def columncount(filename, sheetname):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    return sheet.max_column


def readdata(filename, sheetname, rcount, ccount):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    data = sheet.cell(rcount, ccount).value
    return data


def writedata(filename, sheetname, rcount, ccount, data):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.cell(rcount, ccount).value = data
    workbook.save(filename)


def fillgreencolor(filename, sheetname, rcount, ccount):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    greenfill = PatternFill(start_color='60b212', end_color='60b212', fill_type='solid')
    sheet.cell(rcount, ccount).fill = greenfill
    workbook.save(filename)


def fillredcolor(filename, sheetname, rcount, ccount):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    redfill = PatternFill(start_color='ff0000', end_color='ff0000', fill_type='solid')
    sheet.cell(rcount, ccount).fill = redfill
    workbook.save(filename)
